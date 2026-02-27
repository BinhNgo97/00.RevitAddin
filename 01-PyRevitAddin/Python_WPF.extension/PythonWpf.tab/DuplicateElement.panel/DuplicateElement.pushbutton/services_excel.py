# -*- coding: utf-8 -*-
import clr
import os
clr.AddReference('System.Data')
from System.Data import DataTable

# --- Excel Interop (COM) helpers ---
def _excel_com_open(file_path):
    """Open Excel via COM Interop and return (excel_app, workbook)."""
    try:
        clr.AddReference('Microsoft.Office.Interop.Excel')
        import Microsoft.Office.Interop.Excel as Excel
        from System.Runtime.InteropServices import Marshal

        # Create Excel application
        try:
            excel = Excel.ApplicationClass()
        except:
            excel = Excel.Application()

        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(file_path)
        return excel, wb, Marshal
    except Exception as ex:
        # print("Excel Interop open error: {}".format(str(ex)))
        return None, None, None

def _excel_com_close(excel, workbook, Marshal):
    try:
        if workbook:
            try:
                workbook.Close(False)
            except:
                pass
            try:
                Marshal.FinalReleaseComObject(workbook)
            except:
                pass
        if excel:
            try:
                excel.Quit()
            except:
                pass
            try:
                Marshal.FinalReleaseComObject(excel)
            except:
                pass
    except:
        pass

def _excel_com_list_sheets(file_path):
    """List sheet names using Excel COM Interop."""
    excel, wb, Marshal = _excel_com_open(file_path)
    if not excel or not wb:
        return []
    names = []
    try:
        wsheets = wb.Worksheets
        count = wsheets.Count
        for i in range(1, count + 1):
            try:
                ws = wsheets.Item(i)
                names.append(ws.Name)
            except:
                pass
        # Release worksheets collection
        try:
            Marshal.FinalReleaseComObject(wsheets)
        except:
            pass
    except Exception as ex:
        # print("Excel Interop list sheets error: {}".format(str(ex)))
        pass
    finally:
        _excel_com_close(excel, wb, Marshal)
    return names

def _excel_com_read_rows(file_path, sheet_name):
    """Read all non-empty rows from a sheet using Excel COM Interop. Returns list[list]."""
    excel, wb, Marshal = _excel_com_open(file_path)
    if not excel or not wb:
        return []
    rows = []
    try:
        ws = None
        try:
            ws = wb.Worksheets.Item(sheet_name)
        except:
            # try by index name match
            wsheets = wb.Worksheets
            for i in range(1, wsheets.Count + 1):
                w = wsheets.Item(i)
                if w.Name == sheet_name:
                    ws = w
                    break

        if ws is None:
            print("Excel Interop: sheet not found '{}'".format(sheet_name))
            return rows

        used = ws.UsedRange
        vals = used.Value2

        # Handle empty/None
        if vals is None:
            return rows

        import System
        # System.Array with 2 dimensions; lower bounds may be 1
        if isinstance(vals, System.Array):
            r0 = vals.GetLowerBound(0)
            r1 = vals.GetUpperBound(0)
            c0 = vals.GetLowerBound(1)
            c1 = vals.GetUpperBound(1)
            for r in range(r0, r1 + 1):
                row_vals = []
                for c in range(c0, c1 + 1):
                    try:
                        v = vals.GetValue(r, c)
                    except:
                        v = None
                    row_vals.append("" if v is None else v)
                # trim trailing empties
                while row_vals and (row_vals[-1] is None or str(row_vals[-1]).strip() == ""):
                    row_vals.pop()
                if any(x is not None and str(x).strip() != "" for x in row_vals):
                    rows.append(row_vals)
        else:
            # Single cell or unexpected type
            v = vals
            sval = "" if v is None else str(v)
            if sval.strip() != "":
                rows.append([sval])

        # Release COM objects
        try:
            Marshal.FinalReleaseComObject(used)
        except:
            pass
        try:
            Marshal.FinalReleaseComObject(ws)
        except:
            pass
    except Exception as ex:
        # print("Excel Interop read error: {}".format(str(ex)))
        pass
    finally:
        _excel_com_close(excel, wb, Marshal)
    return rows

# --- Public API ---

def load_excel_sheets(file_path):
    """
    Return list of sheet names.
    Tries openpyxl first; falls back to Excel Interop (COM) if available.
    """
    sheets = []
    if not file_path or not os.path.exists(file_path):
        # print("Excel path invalid")
        return sheets

    # Try openpyxl
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        wb.close()
        # print("Found {} sheets".format(len(sheets)))
        return sheets
    except ImportError:
        # print("openpyxl unavailable in this environment. Trying Excel Interop...")
        pass
    except Exception as ex:
        # print("Error loading workbook with openpyxl: {}".format(str(ex)))
        # print("Trying Excel Interop...")
        pass

    # Fallback: Excel Interop
    try:
        names = _excel_com_list_sheets(file_path)
        if names:
            # print("Found {} sheets via Excel Interop".format(len(names)))
            pass
        return names
    except Exception as ex:
        # print("Excel Interop failed: {}".format(str(ex)))
        return []

def read_sheet_rows(file_path, sheet_name):
    """
    Read all non-empty rows from a sheet.
    Tries openpyxl first; falls back to Excel Interop (COM).
    Returns list[list].
    """
    if not file_path or not os.path.exists(file_path) or not sheet_name:
        return []

    # Try openpyxl
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            # print("Sheet not found: {}".format(sheet_name))
            wb.close()
            return []
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if any(c is not None and str(c).strip() != "" for c in row):
                rows.append([("" if c is None else c) for c in row])
        wb.close()
        return rows
    except ImportError:
        pass
    except Exception as ex:
        # print("Error reading sheet with openpyxl '{}': {}".format(sheet_name, str(ex)))
        pass

    # Fallback: Excel Interop
    try:
        return _excel_com_read_rows(file_path, sheet_name)
    except Exception as ex:
        # print("Excel Interop read failed '{}': {}".format(sheet_name, str(ex)))
        return []

def trim_data_rows(rows):
    if not rows:
        return []
    max_cols = 0
    for row in rows:
        for i in range(len(row) - 1, -1, -1):
            if row[i] is not None and str(row[i]).strip() != "":
                if i + 1 > max_cols:
                    max_cols = i + 1
                break
    if max_cols <= 0:
        return []

    trimmed = []
    for row in rows:
        r = list(row[:max_cols])
        while len(r) < max_cols:
            r.append(None)
        trimmed.append(r)
    return trimmed

def build_datatable_from_rows(rows, add_status=False, check_exists_fn=None, param_comparison_fn=None, name_col_index=None, status_col_index=None):
    dt = DataTable("ExcelData")
    if not rows:
        return dt

    first = list(rows[0] or [])
    has_string_header = any(isinstance(c, basestring) if hasattr(__builtins__, 'basestring') else isinstance(c, str)
                            for c in first if c is not None)
    if has_string_header:
        header = [str(c).strip() if c is not None else "" for c in first]
        data_start = 1
    else:
        header = []
        data_start = 0
    
    # Determine the index of the Name column if not provided
    if name_col_index is None and header:
        # Try to find a column named "Name"
        for i, col_name in enumerate(header):
            if col_name and str(col_name).strip().lower() == "name":
                name_col_index = i
                # print("Found Name column at index {}".format(i))
                break
        
        # If still not found, default to the first column
        if name_col_index is None:
            name_col_index = 0
            # print("No Name column found, defaulting to column 0")
    # print("Using column {} as Name column".format(name_col_index or 0))
    
    # Check if Status column exists in the header
    status_exists = status_col_index is not None
    if not status_exists and header:
        for i, col_name in enumerate(header):
            if col_name and col_name.lower() == "status":
                status_exists = True
                status_col_index = i
                # print("Found existing Status column at index {}".format(i))
                break
                
    action_exists = False
    if header:
        for i, col_name in enumerate(header):
            if col_name and col_name.lower() == "action":
                action_exists = True
                break
                
    # print("Found existing Status column: {} (index: {})".format(status_exists, status_col_index))
    # print("Found existing Action column: {}".format(action_exists))
    # print("Using Name column at index: {}".format(name_col_index or 0))

    # Add columns to DataTable with unique names (only from header or generic Col1..)
    if header and all(h.strip() != "" for h in header):
        seen = {}
        for i, name in enumerate(header):
            base = name if name else "Col{}".format(i + 1)
            n = base
            k = 1
            while n in seen:
                k += 1
                n = "{}_{}".format(base, k)
            seen[n] = True
            while dt.Columns.Contains(n):
                k += 1
                n = "{}_{}".format(base, k)
            dt.Columns.Add(n)
    else:
        max_len = max(len(r) for r in rows)
        for i in range(max_len):
            col_name = "Col{}".format(i + 1)
            k = 1
            while dt.Columns.Contains(col_name):
                k += 1
                col_name = "Col{}_{}".format(i + 1, k)
            dt.Columns.Add(col_name)

    # Ensure a 'Status' column exists (needed by XAML binding [Status])
    if not dt.Columns.Contains("Status"):
        dt.Columns.Add("Status")
        # Do NOT reorder Status to the first position; keep header order intact
        # Removing SetOrdinal(0) prevents shifting data columns and losing the first Excel column

    # Do not add Status or Action columns here; XAML defines those UI columns

    # Process data rows
    for r in range(data_start, len(rows)):
        row = list(rows[r] or [])
        if not any(c is not None and str(c).strip() != "" for c in row):
            continue
        dr = dt.NewRow()
        limit = dt.Columns.Count
        for c in range(min(len(row), limit)):
            v = row[c]
            dr[c] = v if isinstance(v, (int, float)) else ("" if v is None else str(v))

        # Resolve type name
        type_name = ""
        if name_col_index is not None and name_col_index < len(row) and row[name_col_index] is not None:
            type_name = str(row[name_col_index]).strip()

        # Determine status (existing/update/new)
        final_status = ""
        if type_name and callable(check_exists_fn):
            try:
                exists = check_exists_fn(type_name)
                if exists:
                    final_status = "Existing"
                    # Check parameters if comparison function provided
                    if callable(param_comparison_fn):
                        diff_params = param_comparison_fn(type_name, header, row, name_col_index)
                        if diff_params:
                            final_status = "Update"
                            # Store diff list on Status column error (for debugging)
                            try:
                                dr.SetColumnError("Status", "|".join(diff_params))
                            except:
                                pass
                            # Mark each differing parameter cell with an error so the grid can highlight it
                            try:
                                for pname in diff_params:
                                    if pname and dt.Columns.Contains(pname):
                                        dr.SetColumnError(pname, "diff")
                            except:
                                pass
                else:
                    final_status = "New"
            except Exception as ex:
                # print("Error checking type status: {}".format(str(ex)))
                final_status = "Error"
        # Always write computed status (Status column is guaranteed above)
        try:
            dr["Status"] = final_status
        except:
            pass
        dt.Rows.Add(dr)
    return dt

def extract_excel_param_dict(header_row, data_row):
    """Extract parameters from Excel row into name-value dictionary"""
    if not header_row or not data_row:
        return {}
        
    param_dict = {}
    for i in range(1, min(len(header_row), len(data_row))):
        if header_row[i] and header_row[i].strip():
            param_dict[header_row[i].strip()] = data_row[i]
            
    return param_dict
