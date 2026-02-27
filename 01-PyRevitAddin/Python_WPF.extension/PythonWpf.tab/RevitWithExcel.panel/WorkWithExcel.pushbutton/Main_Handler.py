# -*- coding: utf-8 -*-
import clr
clr.AddReference("RevitAPI")
from Autodesk.Revit.DB import *

import openpyxl


class ExcelHandler(object):

    def __init__(self, doc, vm):
        self.doc = doc
        self.vm = vm


    # ------------------------- SELECT ELEMENTS -------------------------
    def select_elements(self, uidoc):
        ref_list = uidoc.Selection.PickObjects(
            ObjectType.Element,
            "Select elements"
        )

        self.vm.selected_elements = [uidoc.Document.GetElement(r.ElementId) for r in ref_list]

        # lấy danh sách parameter duy nhất
        all_params = set()

        for el in self.vm.selected_elements:
            for p in el.Parameters:
                all_params.add(p.Definition.Name)

        self.vm.unique_parameters = sorted(list(all_params))


    # ------------------------- LOAD EXCEL -------------------------
    def set_excel_path(self, path):
        self.vm.excel_path = path


    # ------------------------- REFRESH DATA FROM EXCEL -------------------------
    def refresh_from_excel(self):
        if not self.vm.excel_path:
            return

        wb = openpyxl.load_workbook(self.vm.excel_path)
        ws = wb.active

        headers = [c.value for c in ws[1]]
        data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            d = {}
            for idx, h in enumerate(headers):
                d[h] = row[idx]
            data.append(d)

        self.vm.data_table = data


    # ------------------------- EXPORT DATA TO EXCEL -------------------------
    def export_to_excel(self, selected_params):
        if not self.vm.excel_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active

        headers = ["ID"] + selected_params
        ws.append(headers)

        for el in self.vm.selected_elements:
            row = [el.Id.IntegerValue]

            for p_name in selected_params:
                p = el.LookupParameter(p_name)
                row.append(p.AsString() if p else "")

            ws.append(row)

        wb.save(self.vm.excel_path)
